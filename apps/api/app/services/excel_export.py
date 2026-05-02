from __future__ import annotations

from io import BytesIO
import uuid

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from app.models import (
    ADPEntry,
    ADPSnapshot,
    DraftPick,
    FinalRosterEntry,
    KeeperRecommendation,
    League,
    OptimizerSettings,
    Player,
    Team,
)
from app.services.optimizer import OptimizerInputError, run_optimizer


class ExcelExportError(ValueError):
    """Raised when a keeper export workbook cannot be built."""


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SUBHEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
SECTION_FILL = PatternFill("solid", fgColor="D1FAE5")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14, color="111827")
THIN_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))


def build_keeper_recommendations_workbook(
    session: Session,
    league_id: uuid.UUID,
    *,
    scenario_name: str | None = None,
) -> bytes:
    league = session.get(League, league_id)
    if league is None:
        raise ExcelExportError(f"League {league_id} was not found")

    context = _load_export_context(session, league, scenario_name)
    workbook = Workbook()
    workbook.iso_dates = True
    workbook.remove(workbook.active)

    _write_league_summary(workbook, context)
    _write_drafted_rosters(workbook, context)
    _write_final_rosters(workbook, context)
    _write_adp_input(workbook, context)
    _write_dynamic_keeper_model(workbook, context)
    _write_projected_keepers(workbook, context)
    _write_team_outlooks(workbook, context)
    _write_settings(workbook, context)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _load_export_context(
    session: Session,
    league: League,
    scenario_name: str | None,
) -> dict[str, object]:
    teams = session.exec(select(Team).where(Team.league_id == league.id)).all()
    team_by_id = {team.id: team for team in teams}
    settings = session.exec(
        select(OptimizerSettings).where(OptimizerSettings.league_id == league.id)
    ).first()
    adp_snapshot = session.exec(
        select(ADPSnapshot)
        .where(ADPSnapshot.league_id == league.id)
        .order_by(ADPSnapshot.snapshot_date.desc(), ADPSnapshot.created_at.desc())
    ).first()

    recommendations = _load_recommendations(session, league, scenario_name)
    if not recommendations and adp_snapshot is not None:
        try:
            recommendations = run_optimizer(
                session,
                league.id,
                adp_snapshot_id=adp_snapshot.id,
                scenario_name=scenario_name,
                persist=True,
            )
        except OptimizerInputError as exc:
            raise ExcelExportError(str(exc)) from exc

    draft_picks = session.exec(
        select(DraftPick)
        .where(DraftPick.league_id == league.id, DraftPick.season_year == league.season_year)
        .order_by(DraftPick.overall_pick)
    ).all()
    final_rosters = session.exec(
        select(FinalRosterEntry)
        .where(
            FinalRosterEntry.league_id == league.id,
            FinalRosterEntry.season_year == league.season_year,
        )
    ).all()
    adp_entries = (
        session.exec(
            select(ADPEntry)
            .where(ADPEntry.snapshot_id == adp_snapshot.id)
            .order_by(ADPEntry.adp_pick)
        ).all()
        if adp_snapshot is not None
        else []
    )

    player_ids = {
        row.player_id
        for row in [*draft_picks, *final_rosters, *adp_entries, *recommendations]
        if hasattr(row, "player_id")
    }
    players = (
        {
            player.id: player
            for player in session.exec(select(Player).where(Player.id.in_(player_ids))).all()
        }
        if player_ids
        else {}
    )

    return {
        "league": league,
        "teams": teams,
        "team_by_id": team_by_id,
        "settings": settings,
        "adp_snapshot": adp_snapshot,
        "draft_picks": draft_picks,
        "final_rosters": final_rosters,
        "adp_entries": adp_entries,
        "recommendations": recommendations,
        "players": players,
        "scenario_name": scenario_name,
    }


def _load_recommendations(
    session: Session,
    league: League,
    scenario_name: str | None,
) -> list[KeeperRecommendation]:
    statement = select(KeeperRecommendation).where(KeeperRecommendation.league_id == league.id)
    if scenario_name is not None:
        return session.exec(statement.where(KeeperRecommendation.scenario_name == scenario_name)).all()

    default_rows = session.exec(
        statement.where(KeeperRecommendation.scenario_name == "Default")
    ).all()
    if default_rows:
        return default_rows

    balanced_rows = session.exec(
        statement.where(KeeperRecommendation.scenario_name == "Balanced")
    ).all()
    if balanced_rows:
        return balanced_rows

    return session.exec(statement).all()


def _write_league_summary(workbook: Workbook, context: dict[str, object]) -> None:
    league = context["league"]
    teams = context["teams"]
    recommendations = context["recommendations"]
    assert isinstance(league, League)
    assert isinstance(recommendations, list)

    projected = [
        recommendation
        for recommendation in recommendations
        if isinstance(recommendation, KeeperRecommendation) and recommendation.is_recommended
    ]
    sheet = workbook.create_sheet("League Summary")
    _title(sheet, "League Summary", 1, 1, 6)
    summary_rows = [
        ("League", league.name),
        ("Season", league.season_year),
        ("Scoring Format", league.scoring_format),
        ("Draft Type", league.draft_type),
        ("Teams", len(teams) if isinstance(teams, list) else 0),
        ("Max Keepers", league.max_keepers),
        ("Projected Keepers", len(projected)),
        ("Total Keeper Score", round(sum(row.keeper_score or 0 for row in projected), 3)),
        ("Total Keeper Value", round(sum(row.keeper_value or 0 for row in projected), 3)),
    ]
    _write_key_values(sheet, 3, 1, summary_rows)
    _write_rows(
        sheet,
        3,
        4,
        ["Metric", "Value"],
        [
            ("Drafted Players", len(context["draft_picks"])),
            ("Final Roster Entries", len(context["final_rosters"])),
            ("ADP Entries", len(context["adp_entries"])),
            ("Recommendation Rows", len(recommendations)),
        ],
    )
    _finish_sheet(sheet, widths={1: 22, 2: 28, 4: 24, 5: 18})


def _write_drafted_rosters(workbook: Workbook, context: dict[str, object]) -> None:
    sheet = workbook.create_sheet("Drafted Rosters")
    draft_picks = context["draft_picks"]
    teams = context["team_by_id"]
    players = context["players"]
    assert isinstance(draft_picks, list)
    assert isinstance(teams, dict)
    assert isinstance(players, dict)

    rows = []
    for pick in draft_picks:
        if not isinstance(pick, DraftPick):
            continue
        team = teams.get(pick.team_id)
        player = players.get(pick.player_id)
        rows.append(
            [
                team.name if isinstance(team, Team) else "",
                pick.round,
                pick.overall_pick,
                player.full_name if isinstance(player, Player) else "",
                pick.position,
                player.nfl_team if isinstance(player, Player) else "",
            ]
        )
    _title(sheet, "Drafted Rosters", 1, 1, 6)
    _write_rows(sheet, 3, 1, ["Team", "Round", "Overall Pick", "Player", "Position", "NFL Team"], rows)
    _finish_sheet(sheet, widths={1: 28, 2: 10, 3: 14, 4: 24, 5: 12, 6: 12})


def _write_final_rosters(workbook: Workbook, context: dict[str, object]) -> None:
    sheet = workbook.create_sheet("Final Rosters")
    final_rosters = context["final_rosters"]
    teams = context["team_by_id"]
    players = context["players"]
    assert isinstance(final_rosters, list)
    assert isinstance(teams, dict)
    assert isinstance(players, dict)

    rows = []
    for entry in final_rosters:
        if not isinstance(entry, FinalRosterEntry):
            continue
        team = teams.get(entry.team_id)
        player = players.get(entry.player_id)
        rows.append(
            [
                team.name if isinstance(team, Team) else "",
                player.full_name if isinstance(player, Player) else "",
                entry.position,
                player.nfl_team if isinstance(player, Player) else "",
                entry.roster_status,
            ]
        )
    _title(sheet, "Final Rosters", 1, 1, 5)
    _write_rows(sheet, 3, 1, ["Team", "Player", "Position", "NFL Team", "Roster Status"], rows)
    _finish_sheet(sheet, widths={1: 28, 2: 24, 3: 12, 4: 12, 5: 18})


def _write_adp_input(workbook: Workbook, context: dict[str, object]) -> None:
    sheet = workbook.create_sheet("ADP Input")
    snapshot = context["adp_snapshot"]
    adp_entries = context["adp_entries"]
    players = context["players"]
    assert isinstance(adp_entries, list)
    assert isinstance(players, dict)

    _title(sheet, "ADP Input", 1, 1, 7)
    if isinstance(snapshot, ADPSnapshot):
        _write_key_values(
            sheet,
            3,
            1,
            [
                ("Snapshot", snapshot.name),
                ("Source", snapshot.source),
                ("Format", snapshot.format_type),
                ("Snapshot Date", snapshot.snapshot_date),
            ],
        )
        start_row = 9
    else:
        sheet.cell(3, 1, "No ADP snapshot available")
        start_row = 5

    rows = []
    for entry in adp_entries:
        if not isinstance(entry, ADPEntry):
            continue
        player = players.get(entry.player_id)
        rows.append(
            [
                player.full_name if isinstance(player, Player) else "",
                entry.position,
                entry.adp_pick,
                entry.adp_round,
                player.nfl_team if isinstance(player, Player) else "",
                entry.source_note,
            ]
        )
    _write_rows(
        sheet,
        start_row,
        1,
        ["Player", "Position", "ADP Pick", "ADP Round", "NFL Team", "Source Note"],
        rows,
    )
    _finish_sheet(sheet, widths={1: 24, 2: 12, 3: 12, 4: 12, 5: 12, 6: 28})


def _write_dynamic_keeper_model(workbook: Workbook, context: dict[str, object]) -> None:
    sheet = workbook.create_sheet("Dynamic Keeper Model")
    recommendations = context["recommendations"]
    teams = context["team_by_id"]
    players = context["players"]
    assert isinstance(recommendations, list)
    assert isinstance(teams, dict)
    assert isinstance(players, dict)

    headers = [
        "Team",
        "Player",
        "Position",
        "Roster Status",
        "Keeper Cost Pick",
        "Keeper Cost Round",
        "ADP Pick",
        "ADP Round",
        "Keeper Value",
        "Keeper Score",
        "Eligible",
        "Projected Keeper",
        "Reason",
    ]
    rows = []
    for recommendation in recommendations:
        if not isinstance(recommendation, KeeperRecommendation):
            continue
        team = teams.get(recommendation.team_id)
        player = players.get(recommendation.player_id)
        rows.append(
            [
                team.name if isinstance(team, Team) else "",
                player.full_name if isinstance(player, Player) else "",
                player.position if isinstance(player, Player) else "",
                "",
                recommendation.keeper_cost_pick,
                recommendation.keeper_cost_round,
                recommendation.adp_pick,
                recommendation.adp_round,
                None,
                recommendation.keeper_score,
                recommendation.is_eligible,
                recommendation.is_recommended,
                recommendation.reason,
            ]
        )

    _title(sheet, "Dynamic Keeper Model", 1, 1, len(headers))
    _write_rows(sheet, 3, 1, headers, rows)
    for row_index in range(4, 4 + len(rows)):
        sheet.cell(row_index, 9, f'=IF(OR(E{row_index}="",G{row_index}=""),"",E{row_index}-G{row_index})')
    _finish_sheet(
        sheet,
        widths={
            1: 28,
            2: 24,
            3: 12,
            4: 16,
            5: 18,
            6: 18,
            7: 12,
            8: 12,
            9: 14,
            10: 14,
            11: 12,
            12: 18,
            13: 32,
        },
    )


def _write_projected_keepers(workbook: Workbook, context: dict[str, object]) -> None:
    sheet = workbook.create_sheet("Projected Keepers")
    recommendations = context["recommendations"]
    teams = context["team_by_id"]
    players = context["players"]
    assert isinstance(recommendations, list)
    assert isinstance(teams, dict)
    assert isinstance(players, dict)

    rows = []
    for recommendation in recommendations:
        if not isinstance(recommendation, KeeperRecommendation) or not recommendation.is_recommended:
            continue
        team = teams.get(recommendation.team_id)
        player = players.get(recommendation.player_id)
        rows.append(
            [
                team.name if isinstance(team, Team) else "",
                player.full_name if isinstance(player, Player) else "",
                player.position if isinstance(player, Player) else "",
                recommendation.keeper_cost_pick,
                recommendation.keeper_cost_round,
                recommendation.adp_pick,
                recommendation.keeper_value,
                recommendation.keeper_score,
                recommendation.reason,
            ]
        )
    _title(sheet, "Projected Keepers", 1, 1, 9)
    _write_rows(
        sheet,
        3,
        1,
        [
            "Team",
            "Player",
            "Position",
            "Cost Pick",
            "Cost Round",
            "ADP Pick",
            "Keeper Value",
            "Keeper Score",
            "Reason",
        ],
        rows,
    )
    _finish_sheet(sheet, widths={1: 28, 2: 24, 3: 12, 4: 12, 5: 12, 6: 12, 7: 14, 8: 14, 9: 34})


def _write_team_outlooks(workbook: Workbook, context: dict[str, object]) -> None:
    sheet = workbook.create_sheet("Team Outlooks")
    teams = context["teams"]
    recommendations = context["recommendations"]
    players = context["players"]
    assert isinstance(teams, list)
    assert isinstance(recommendations, list)
    assert isinstance(players, dict)

    projected_by_team: dict[uuid.UUID, list[KeeperRecommendation]] = {}
    for recommendation in recommendations:
        if isinstance(recommendation, KeeperRecommendation) and recommendation.is_recommended:
            projected_by_team.setdefault(recommendation.team_id, []).append(recommendation)

    rows = []
    for team in teams:
        if not isinstance(team, Team):
            continue
        keepers = projected_by_team.get(team.id, [])
        keeper_names = [
            players[row.player_id].full_name
            for row in keepers
            if row.player_id in players and isinstance(players[row.player_id], Player)
        ]
        picks = [_format_pick(row) for row in keepers]
        total_score = round(sum(row.keeper_score or 0 for row in keepers), 3)
        rows.append(
            [
                team.name,
                team.draft_slot,
                len(keepers),
                ", ".join(keeper_names),
                ", ".join(picks),
                total_score,
                _outlook_note(len(keepers), total_score, picks),
            ]
        )
    _title(sheet, "Team Outlooks", 1, 1, 7)
    _write_rows(
        sheet,
        3,
        1,
        [
            "Team",
            "Draft Slot",
            "Projected Keepers",
            "Keeper Names",
            "Picks Forfeited",
            "Total Keeper Score",
            "Strategic Notes",
        ],
        rows,
    )
    _finish_sheet(sheet, widths={1: 28, 2: 12, 3: 18, 4: 36, 5: 28, 6: 18, 7: 46})


def _write_settings(workbook: Workbook, context: dict[str, object]) -> None:
    sheet = workbook.create_sheet("Settings")
    league = context["league"]
    settings = context["settings"]
    scenario_name = context["scenario_name"]
    assert isinstance(league, League)

    _title(sheet, "Settings", 1, 1, 4)
    rows = [
        ("League Max Keepers", league.max_keepers),
        ("Max Keepers Per Position", league.max_keepers_per_position),
        ("Max QB Keepers", league.max_qb_keepers),
        ("Export Scenario", scenario_name or "Default/Balanced"),
    ]
    if isinstance(settings, OptimizerSettings):
        rows.extend(
            [
                ("Settings Name", settings.name),
                ("Minimum Keeper Value", settings.minimum_keeper_value),
                ("Minimum Keeper Score", settings.minimum_keeper_score),
                ("Max ADP Cap", settings.max_adp_cap),
                ("QB Weight", settings.qb_weight),
                ("RB Weight", settings.rb_weight),
                ("WR Weight", settings.wr_weight),
                ("TE Weight", settings.te_weight),
                ("K Weight", settings.k_weight),
                ("DEF Weight", settings.def_weight),
                ("QB Max ADP", settings.qb_max_adp),
                ("Elite QB Cutoff", settings.elite_qb_cutoff),
                ("Talent Anchor", settings.talent_anchor),
                ("Talent Divisor", settings.talent_divisor),
                ("Starter Status Bonus", settings.starter_status_bonus),
                ("Bench Status Bonus", settings.bench_status_bonus),
                ("IR Status Bonus", settings.ir_status_bonus),
                ("Draft Slot Bonus Enabled", settings.enable_draft_slot_bonus),
                ("QB Scarcity Bonus Enabled", settings.enable_qb_scarcity_bonus),
            ]
        )
    _write_rows(sheet, 3, 1, ["Setting", "Value"], rows)
    _finish_sheet(sheet, widths={1: 30, 2: 20})


def _write_key_values(sheet, start_row: int, start_col: int, rows: list[tuple[object, object]]) -> None:
    _write_rows(sheet, start_row, start_col, ["Field", "Value"], rows)


def _write_rows(sheet, start_row: int, start_col: int, headers: list[str], rows: list[object]) -> None:
    for offset, header in enumerate(headers):
        cell = sheet.cell(start_row, start_col + offset, header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    if not rows:
        sheet.cell(start_row + 1, start_col, "No rows")
        return

    for row_offset, row in enumerate(rows, start=1):
        values = list(row) if isinstance(row, (list, tuple)) else [row]
        for col_offset, value in enumerate(values):
            cell = sheet.cell(start_row + row_offset, start_col + col_offset, value)
            cell.border = THIN_BORDER
            if isinstance(value, bool):
                cell.value = "Yes" if value else "No"


def _title(sheet, title: str, row: int, start_col: int, end_col: int) -> None:
    sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = sheet.cell(row, start_col, title)
    cell.font = TITLE_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left")


def _finish_sheet(sheet, widths: dict[int, int]) -> None:
    sheet.freeze_panes = "A4"
    for col_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "left",
                vertical="top",
                wrap_text=True,
            )
    for cell in sheet[3]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = WHITE_FONT
    sheet.sheet_view.showGridLines = False


def _format_pick(recommendation: KeeperRecommendation) -> str:
    if recommendation.keeper_cost_pick is None:
        return ""
    if recommendation.keeper_cost_round is None:
        return f"Pick {recommendation.keeper_cost_pick:g}"
    return f"R{recommendation.keeper_cost_round:g} / Pick {recommendation.keeper_cost_pick:g}"


def _outlook_note(keeper_count: int, total_score: float, picks: list[str]) -> str:
    if keeper_count == 0:
        return "No projected keepers; draft flexibility preserved."
    if keeper_count >= 4:
        return f"Full keeper set worth {total_score:g}; confirm pick costs before deadline."
    if any("R1" in pick or "R2" in pick for pick in picks):
        return f"{keeper_count} keeper(s) with premium pick cost; compare against draft pool."
    return f"{keeper_count} keeper(s) selected with manageable pick cost."
