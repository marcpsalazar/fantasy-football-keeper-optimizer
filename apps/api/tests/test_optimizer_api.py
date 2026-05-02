from collections.abc import Generator
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel, Session, create_engine

import app.models  # noqa: F401
from app.db.session import get_session
from app.main import create_app


@pytest.fixture
def client() -> Generator[TestClient, None, None]:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)

    def override_get_session() -> Generator[Session, None, None]:
        with Session(engine) as session:
            yield session

    app = create_app()
    app.dependency_overrides[get_session] = override_get_session

    with TestClient(app) as test_client:
        yield test_client

    app.dependency_overrides.clear()


def test_optimizer_run_endpoint_returns_frontend_table(client: TestClient) -> None:
    league_response = client.post(
        "/api/leagues",
        json={
            "name": "Endpoint League",
            "season_year": 2026,
            "scoring_format": "superflex",
        },
    )
    assert league_response.status_code == 201
    league_id = league_response.json()["id"]

    team_response = client.post(
        f"/api/leagues/{league_id}/teams",
        json={"name": "Endpoint Team", "draft_slot": 1},
    )
    assert team_response.status_code == 201

    other_team_response = client.post(
        f"/api/leagues/{league_id}/teams",
        json={"name": "Other Team", "draft_slot": 2},
    )
    assert other_team_response.status_code == 201
    other_team_id = other_team_response.json()["id"]

    league_update_response = client.patch(
        f"/api/leagues/{league_id}",
        json={"keeper_rules": {"cost": "same-team draft pick or ADP"}},
    )
    assert league_update_response.status_code == 200
    assert league_update_response.json()["keeper_rules"]["cost"] == "same-team draft pick or ADP"

    team_update_response = client.patch(
        f"/api/teams/{other_team_id}",
        json={"owner_name": "Other Owner"},
    )
    assert team_update_response.status_code == 200
    assert team_update_response.json()["owner_name"] == "Other Owner"

    invalid_draft_preview = client.post(
        f"/api/leagues/{league_id}/draft-results/preview",
        content=(
            "team,round,overall_pick,player,position\n"
            "Endpoint Team,5,50,Value WR,WR\n"
            "Ghost Team,5,50,Copy WR,WR\n"
            "Endpoint Team,0,nope,Bad Player,FLEX\n"
        ),
        headers={"content-type": "text/csv"},
    )
    assert invalid_draft_preview.status_code == 200
    invalid_draft_payload = invalid_draft_preview.json()
    assert invalid_draft_payload["valid"] is False
    assert invalid_draft_payload["error_count"] == 4
    assert invalid_draft_payload["warning_count"] == 1
    assert any(error["field"] == "overall_pick" for error in invalid_draft_payload["errors"])
    assert invalid_draft_payload["warnings"][0]["field"] == "team"

    valid_draft_preview = client.post(
        f"/api/leagues/{league_id}/draft-results/preview",
        content=(
            "team,round,overall_pick,player,position\n"
            "Endpoint Team,5,50,Value WR,WR\n"
            "Endpoint Team,8,90,Value RB,RB\n"
        ),
        headers={"content-type": "text/csv"},
    )
    assert valid_draft_preview.status_code == 200
    assert valid_draft_preview.json()["valid"] is True
    assert valid_draft_preview.json()["valid_rows"] == 2

    draft_response = client.post(
        f"/api/leagues/{league_id}/draft-results/import",
        content=(
            "team,round,overall_pick,player,position\n"
            "Endpoint Team,5,50,Value WR,WR\n"
            "Endpoint Team,8,90,Value RB,RB\n"
            "Endpoint Team,2,20,Market QB,QB\n"
            "Other Team,1,1,Other WR,WR\n"
        ),
        headers={"content-type": "text/csv"},
    )
    assert draft_response.status_code == 200
    assert draft_response.json()["imported"] == 4
    draft_results_response = client.get(f"/api/leagues/{league_id}/draft-results")
    assert draft_results_response.status_code == 200
    assert draft_results_response.json()["count"] == 4
    assert draft_results_response.json()["rows"][0]["player_name"] == "Other WR"

    roster_response = client.post(
        f"/api/leagues/{league_id}/final-rosters/import",
        content=(
            "team,player,position,roster_status\n"
            "Endpoint Team,Value WR,WR,Starter\n"
            "Endpoint Team,Value RB,RB,Starter\n"
            "Endpoint Team,Market QB,QB,Starter\n"
        ),
        headers={"content-type": "text/csv"},
    )
    assert roster_response.status_code == 200
    assert roster_response.json()["count"] == 3
    roster_results_response = client.get(f"/api/leagues/{league_id}/final-rosters")
    assert roster_results_response.status_code == 200
    assert roster_results_response.json()["count"] == 3
    assert roster_results_response.json()["rows"][0]["acquired_via"] == "Drafted"

    invalid_roster_preview = client.post(
        f"/api/leagues/{league_id}/final-rosters/preview",
        content=(
            "team,player,position,roster_status\n"
            "Endpoint Team,Value WR,WR,Starter\n"
            "Endpoint Team,Value WR,WR,Bench\n"
            "Endpoint Team,Bad Position,FLEX,Bench\n"
        ),
        headers={"content-type": "text/csv"},
    )
    assert invalid_roster_preview.status_code == 200
    assert invalid_roster_preview.json()["valid"] is False
    assert invalid_roster_preview.json()["error_count"] == 2

    adp_response = client.post(
        f"/api/leagues/{league_id}/adp/import",
        content=(
            "player,position,adp_pick,source,snapshot_date,format\n"
            "Value WR,WR,15,Endpoint ADP,2026-05-01,superflex\n"
            "Value RB,RB,40,Endpoint ADP,2026-05-01,superflex\n"
            "Market QB,QB,20,Endpoint ADP,2026-05-01,superflex\n"
        ),
        headers={"content-type": "text/csv"},
    )
    assert adp_response.status_code == 200
    assert adp_response.json()["count"] == 3
    snapshots_response = client.get(f"/api/leagues/{league_id}/adp-snapshots")
    assert snapshots_response.status_code == 200
    assert snapshots_response.json()["count"] == 1
    snapshot_id = snapshots_response.json()["rows"][0]["id"]
    snapshot_response = client.get(f"/api/adp-snapshots/{snapshot_id}")
    assert snapshot_response.status_code == 200
    assert snapshot_response.json()["snapshot"]["entry_count"] == 3
    assert snapshot_response.json()["rows"][0]["player_name"] == "Value WR"

    invalid_adp_preview = client.post(
        f"/api/leagues/{league_id}/adp/preview",
        content=(
            "player,position,adp_pick,source,snapshot_date\n"
            "Value WR,WR,15,Endpoint ADP,2026-05-01\n"
            "Value WR,WR,16,Endpoint ADP,2026-05-01\n"
            "Bad ADP,FLEX,-1,Endpoint ADP,not-a-date\n"
        ),
        headers={"content-type": "text/csv"},
    )
    assert invalid_adp_preview.status_code == 200
    assert invalid_adp_preview.json()["valid"] is False
    assert invalid_adp_preview.json()["error_count"] == 4

    settings_read_response = client.get(f"/api/leagues/{league_id}/optimizer/settings")
    assert settings_read_response.status_code == 200

    settings_response = client.patch(
        f"/api/leagues/{league_id}/optimizer/settings",
        json={
            "minimum_keeper_value": 1,
            "minimum_keeper_score": 0,
            "max_keepers": 4,
            "max_keepers_per_position": 2,
            "max_qb_keepers": 1,
        },
    )
    assert settings_response.status_code == 200

    run_response = client.post(f"/api/leagues/{league_id}/optimizer/run", json={})
    assert run_response.status_code == 200
    payload = run_response.json()

    assert payload["count"] == 3
    assert payload["selected_count"] == 2
    assert {"team_name", "player_name", "keeper_score", "is_recommended"}.issubset(
        payload["columns"]
    )

    selected_names = {
        row["player_name"] for row in payload["rows"] if row["is_recommended"]
    }
    assert selected_names == {"Value WR", "Value RB"}

    qb_row = next(row for row in payload["rows"] if row["player_name"] == "Market QB")
    assert qb_row["is_recommended"] is False
    assert qb_row["reason"] == "Keeper value below minimum"
    value_rb_row = next(row for row in payload["rows"] if row["player_name"] == "Value RB")

    force_response = client.put(
        f"/api/leagues/{league_id}/manual-overrides",
        json={
            "team_id": qb_row["team_id"],
            "player_id": qb_row["player_id"],
            "override_type": "force_keep",
        },
    )
    assert force_response.status_code == 200
    assert force_response.json()["override_type"] == "force_keep"
    exclude_response = client.put(
        f"/api/leagues/{league_id}/manual-overrides",
        json={
            "team_id": value_rb_row["team_id"],
            "player_id": value_rb_row["player_id"],
            "override_type": "exclude",
            "notes": "Testing override flow",
        },
    )
    assert exclude_response.status_code == 200
    assert exclude_response.json()["notes"] == "Testing override flow"
    overrides_response = client.get(f"/api/leagues/{league_id}/manual-overrides")
    assert overrides_response.status_code == 200
    assert overrides_response.json()["count"] == 2

    rerun_response = client.post(f"/api/leagues/{league_id}/optimizer/run", json={})
    assert rerun_response.status_code == 200
    rerun_payload = rerun_response.json()
    overridden_selected = {
        row["player_name"] for row in rerun_payload["rows"] if row["is_recommended"]
    }
    assert overridden_selected == {"Value WR", "Market QB"}
    overridden_rb = next(row for row in rerun_payload["rows"] if row["player_name"] == "Value RB")
    assert overridden_rb["manual_override"] == "exclude"
    assert overridden_rb["reason"] == "Manual override excluded player"

    clear_response = client.put(
        f"/api/leagues/{league_id}/manual-overrides",
        json={
            "team_id": qb_row["team_id"],
            "player_id": qb_row["player_id"],
            "override_type": "auto",
        },
    )
    assert clear_response.status_code == 200
    assert clear_response.json()["override_type"] == "auto"

    results_response = client.get(f"/api/leagues/{league_id}/optimizer/results")
    assert results_response.status_code == 200
    results_payload = results_response.json()
    assert results_payload["count"] == 3
    assert results_payload["selected_count"] == 2

    draft_impact_response = client.get(f"/api/leagues/{league_id}/draft-impact?rounds=30")
    assert draft_impact_response.status_code == 200
    draft_impact_payload = draft_impact_response.json()
    assert draft_impact_payload["forfeited_count"] == 2
    forfeited = [row for row in draft_impact_payload["rows"] if row["status"] == "Forfeited"]
    assert {row["keeper_player"] for row in forfeited} == {"Value WR", "Market QB"}

    clear_rb_response = client.put(
        f"/api/leagues/{league_id}/manual-overrides",
        json={
            "team_id": value_rb_row["team_id"],
            "player_id": value_rb_row["player_id"],
            "override_type": "auto",
        },
    )
    assert clear_rb_response.status_code == 200
    restore_response = client.post(f"/api/leagues/{league_id}/optimizer/run", json={})
    assert restore_response.status_code == 200
    restored_selected = {
        row["player_name"] for row in restore_response.json()["rows"] if row["is_recommended"]
    }
    assert restored_selected == {"Value WR", "Value RB"}

    scenarios_response = client.post(
        f"/api/leagues/{league_id}/optimizer/scenarios",
        json={"persist": False},
    )
    assert scenarios_response.status_code == 200
    scenarios_payload = scenarios_response.json()
    assert scenarios_payload["scenario_names"] == [
        "Pure Value",
        "Balanced",
        "Superflex Heavy",
        "Win Now",
        "Rebuild",
    ]
    assert scenarios_payload["count"] == 5
    assert "Endpoint Team" in scenarios_payload["team_names"]

    balanced = next(
        scenario
        for scenario in scenarios_payload["scenarios"]
        if scenario["scenario_name"] == "Balanced"
    )
    endpoint_team = next(
        team
        for team in balanced["teams"]
        if team["team_name"] == "Endpoint Team"
    )
    assert endpoint_team["total_keeper_score"] > 0
    assert endpoint_team["picks_forfeited"] == ["R5 / Pick 50", "R8 / Pick 90"]
    assert [keeper["player_name"] for keeper in endpoint_team["selected_keepers"]] == [
        "Value WR",
        "Value RB",
    ]
    assert endpoint_team["strategic_notes"]

    csv_export_response = client.get(
        f"/api/leagues/{league_id}/exports/keeper-recommendations.csv"
    )
    assert csv_export_response.status_code == 200
    assert "Value WR" in csv_export_response.text
    assert "Value RB" in csv_export_response.text

    pdf_export_response = client.get(f"/api/leagues/{league_id}/exports/team-outlooks.pdf")
    assert pdf_export_response.status_code == 200
    assert pdf_export_response.headers["content-type"] == "application/pdf"
    assert pdf_export_response.content.startswith(b"%PDF-1.4")

    export_response = client.get(
        f"/api/leagues/{league_id}/exports/keeper-recommendations.xlsx"
    )
    assert export_response.status_code == 200
    assert export_response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(export_response.content), data_only=False)
    assert workbook.sheetnames == [
        "League Summary",
        "Drafted Rosters",
        "Final Rosters",
        "ADP Input",
        "Dynamic Keeper Model",
        "Projected Keepers",
        "Team Outlooks",
        "Settings",
    ]
    assert all(not worksheet.tables for worksheet in workbook.worksheets)
    assert workbook["League Summary"]["A1"].value == "League Summary"
    assert workbook["Projected Keepers"]["A3"].value == "Team"
    assert workbook["Dynamic Keeper Model"]["I4"].value == '=IF(OR(E4="",G4=""),"",E4-G4)'
    projected_players = [
        workbook["Projected Keepers"].cell(row=row, column=2).value
        for row in range(4, workbook["Projected Keepers"].max_row + 1)
    ]
    assert {"Value WR", "Value RB"}.issubset(set(projected_players))
